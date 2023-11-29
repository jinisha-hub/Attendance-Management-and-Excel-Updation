from contextlib import nullcontext
import pyrebase
import email
from app1.models import Section
from app1.models import Year
from multiprocessing import context
from pyexpat.errors import messages
from django.shortcuts import render,HttpResponse,redirect
from django.contrib.auth.models import User
from django.contrib.auth import authenticate,login,logout
from django.contrib.auth.decorators import login_required
from django.views import View
import openpyxl
from requests import request
from app1.models import Branch, Email, Home, Section, Year
from app1.models import Attendance
from datetime import datetime
#from django.contrib.sessions.models import Session
#from django.utils import timezone
from openpyxl.workbook import workbook
from openpyxl import load_workbook


config={
   'apiKey': "AIzaSyD8CDI7MyS1dmU2P0dlNPnLO9a6DwnMiJU",
   'databaseURL' : "",
    'authDomain': "feedback-39ef2.firebaseapp.com",
    'projectId': "feedback-39ef2",
    'storageBucket': "feedback-39ef2.appspot.com",
    'messagingSenderId': "148879615339",
    'appId': "1:148879615339:web:b4c372fbae37a9ab3f982e",
    'measurementId': "G-8Y6TW1CQTB"
}
firebase=pyrebase.initialize_app(config)

auth= firebase.auth()

# Create your views here.
@login_required(login_url='login')
def HomePage(request):
    #current_user = request.user
    #user_username = current_user.username
    #print(user_username)
    #context = {'user_username': user_username}
    if request.method=="POST":
        teacher=request.POST.get('teacher')
        score=request.POST.get('score')
        feedback=request.POST.get('feedback')
        if not teacher or not feedback or not score:
          return HttpResponse("Enter the details!!!")
        else:
         obj=Home (teacher=teacher,feedback=feedback,date=datetime.now(),score=score)
         obj.save()
        #messages.success(request,'Your Attendance has been Recorded')
        return redirect('thanks')
    return render (request,'home.html')

def SignupPage(request):
    if request.method=='POST':
        uname=request.POST.get('username')
        email=request.POST.get('email')
        pass1=request.POST.get('password1')
        pass2=request.POST.get('password2')

        if pass1!=pass2:
            return HttpResponse("Your password and confrom password are not Same!!!")
        else:

            my_user=User.objects.create_user(uname,email,pass1)
            my_user.save()
            return redirect('login')
        



    return render (request,'signup.html')

def LoginPage(request):
   
    if request.method=='POST':
        username=request.POST.get('username')
        pass1=request.POST.get('pass')
        user=authenticate(request,username=username,password=pass1)
        if user is not None:
            login(request,user)
            return redirect('attendance')
        else:
            return HttpResponse ("Username or Password is incorrect!!!")
         
    return render (request,'login.html')

def LogoutPage(request):
    logout(request)
    return redirect('login')


def attendance(request):
    if request.method == 'GET':
     
     
     branchid = request.GET.get('branch', None)
     yearid = request.GET.get('year', None)
     sectionid = request.GET.get('section', None)
     emailid = request.GET.get('email', None)
     year = None
     section = None
     email=None
     
     if branchid:
        getbranch = Branch.objects.get(name=branchid)
        year = Year.objects.filter(branch=getbranch)
     if yearid:
        getyear = Year.objects.get(name=yearid)
        section = Section.objects.filter(year=getyear)
     if sectionid:
        getsection = Section.objects.get(name=sectionid)
        section = Email.objects.filter(section=getsection)
        
     if emailid:
        getemail = Email.objects.get(name=yearid)

     return render(request, 'attendance.html', locals())


@login_required(login_url='login')
def attendance(request):
    if request.method == 'POST':
     
          teacher=request.POST.get('teacher')
          branch=request.POST.get('branch')
          year=request.POST.get('year')
          section=request.POST.get('section')
          email=request.POST.get('email')
          if not branch or not year  or not section or not email:
            return HttpResponse("Enter the details!!!")
          else:   
            
           objj=Attendance (date=datetime.today(),branch=branch,year=year,section=section,email=email,teacher=teacher)
           print(teacher)
           if teacher == "0":
                
                print(teacher)
                file_path = "teacher1.xlsx"
           elif teacher == "1":
                file_path = "teacher2.xlsx" 
           wb=openpyxl.load_workbook(file_path)
           
           sheet_names = ['Sheet2', 'Sheet1','Sheet4']
           sheets_dict = {sheet_name: wb[sheet_name] for sheet_name in sheet_names}
       
           date = datetime.now().date()
           date_str = date.strftime("%Y-%m-%d")
           

          for sheet_names, ws in sheets_dict.items(): 
           for j in range(2,100):
            if (ws.cell(row=1,column=j).value) == date_str :
             print('hey')
             for i in range(2,15):
               first=ws.cell(row=i,column=1).value
               print(first)
               if first == email:
                ws.cell(row=i,column=j).value='present'
                print('present')
               if teacher == "teacher1":
                 print(teacher)
                 file_path = "teacher1.xlsx"
               elif teacher == "teacher2":
                 file_path = "teacher2.xlsx"  
             wb.save(file_path)
             
             objj.save()
           

          return redirect('home')
            
    return render (request,'attendance.html')

def Thanks(request):
   
    return render (request,'thanks.html')

def get_years(request):
    branch_name = request.GET['branch_name']
    get_branch = Branch.objects.get(id=branch_name)
    year = Year.objects.filter(branch=get_branch)
    return render(request, 'get-years.html', locals())

def get_sections(request):
    year_name = request.GET['year_name']
    get_year = Year.objects.get(name=year_name)
    section = Section.objects.filter(year=get_year)
    return render(request, 'get-sections.html', locals())
 
def get_emails(request):
    section_name = request.GET['section_name']
    get_section = Section.objects.get(name=section_name)
    email = Email.objects.filter(section=get_section)
    return render(request, 'get-emails.html', locals()) 







